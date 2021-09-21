from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

wb = load_workbook(filename = 'lemmi-export.xlsx')
ws = wb.active

lemmi = []

for row in range(2, 15480):
    lemmi.append(ws["B" + str(row)].value)


for lemma in lemmi:
    diminutivi = re.match(".*in[aeiou]|.*ett[aeiou]|.*ucci[aeiou]|.*ell[aeiou]", lemma)
    if diminutivi:
        with open('esclusioni_diminutivi.txt', 'a') as file:
            file.write(diminutivi.group() + "\n")